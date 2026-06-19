"""ETL pipeline for the UCI Online Retail II workbook."""

from __future__ import annotations

import argparse
import csv
import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

import duckdb
from openpyxl import load_workbook


HEADER_MAP = {
    "Invoice": "invoice_no",
    "StockCode": "stock_code",
    "Description": "description",
    "Quantity": "quantity",
    "InvoiceDate": "invoice_timestamp",
    "Price": "unit_price",
    "Customer ID": "customer_id",
    "Country": "country",
}
STAGING_FIELDS = [
    "source_period",
    "invoice_no",
    "stock_code",
    "description",
    "quantity",
    "invoice_timestamp",
    "unit_price",
    "customer_id",
    "country",
]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _identifier(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _timestamp(value: Any) -> str:
    if isinstance(value, datetime):
        timestamp = value
    elif isinstance(value, date):
        timestamp = datetime.combine(value, datetime.min.time())
    else:
        timestamp = datetime.fromisoformat(str(value).strip())
    return timestamp.isoformat(sep=" ", timespec="seconds")


def _normalise_row(record: dict[str, Any], source_period: str) -> dict[str, Any]:
    quantity = int(record["Quantity"])
    unit_price = float(record["Price"])
    return {
        "source_period": source_period,
        "invoice_no": _identifier(record["Invoice"]),
        "stock_code": _identifier(record["StockCode"]),
        "description": _text(record["Description"]),
        "quantity": quantity,
        "invoice_timestamp": _timestamp(record["InvoiceDate"]),
        "unit_price": unit_price,
        "customer_id": _identifier(record["Customer ID"]),
        "country": _text(record["Country"]),
    }


def workbook_to_staging_csv(
    workbook_path: str | Path,
    staging_path: str | Path,
    quarantine_path: str | Path,
) -> tuple[int, int]:
    """Stream workbook rows into a normalized CSV and quarantine invalid rows."""
    workbook_path = Path(workbook_path)
    staging_path = Path(staging_path)
    quarantine_path = Path(quarantine_path)
    staging_path.parent.mkdir(parents=True, exist_ok=True)
    valid_rows = 0
    quarantined_rows = 0

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        with (
            staging_path.open("w", newline="", encoding="utf-8") as staging_file,
            quarantine_path.open("w", newline="", encoding="utf-8")
            as quarantine_file,
        ):
            staging_writer = csv.DictWriter(staging_file, fieldnames=STAGING_FIELDS)
            quarantine_writer = csv.DictWriter(
                quarantine_file,
                fieldnames=["source_period", "source_row", "error", "raw_values"],
            )
            staging_writer.writeheader()
            quarantine_writer.writeheader()

            for worksheet in workbook.worksheets:
                rows = worksheet.iter_rows(values_only=True)
                headers = next(rows)
                if set(headers) != set(HEADER_MAP):
                    raise ValueError(
                        f"Unexpected columns in sheet '{worksheet.title}': {headers}"
                    )

                for source_row, values in enumerate(rows, start=2):
                    record = dict(zip(headers, values))
                    try:
                        staging_writer.writerow(
                            _normalise_row(record, worksheet.title)
                        )
                        valid_rows += 1
                    except (KeyError, TypeError, ValueError) as error:
                        quarantine_writer.writerow(
                            {
                                "source_period": worksheet.title,
                                "source_row": source_row,
                                "error": str(error),
                                "raw_values": repr(values),
                            }
                        )
                        quarantined_rows += 1
    finally:
        workbook.close()

    return valid_rows, quarantined_rows


def _write_curated_parquet(staging_path: Path, output_path: Path) -> None:
    column_types = {
        "source_period": "VARCHAR",
        "invoice_no": "VARCHAR",
        "stock_code": "VARCHAR",
        "description": "VARCHAR",
        "quantity": "INTEGER",
        "invoice_timestamp": "TIMESTAMP",
        "unit_price": "DOUBLE",
        "customer_id": "BIGINT",
        "country": "VARCHAR",
    }

    with duckdb.connect() as connection:
        source = connection.read_csv(
            str(staging_path),
            header=True,
            columns=column_types,
        )
        source.project(
            """
            row_number() OVER ()::BIGINT AS transaction_line_id,
            source_period,
            invoice_no,
            stock_code,
            description,
            quantity,
            invoice_timestamp,
            unit_price,
            customer_id,
            country,
            starts_with(upper(invoice_no), 'C') AS is_cancelled,
            round(quantity * unit_price, 2) AS line_revenue,
            (
                NOT starts_with(upper(invoice_no), 'C')
                AND quantity > 0
                AND unit_price > 0
            ) AS is_valid_sale
            """
        ).write_parquet(
            str(output_path),
            compression="zstd",
            overwrite=True,
            row_group_size=100_000,
        )


def retail_quality_summary(parquet_path: str | Path) -> dict[str, Any]:
    parquet_path = Path(parquet_path).resolve()
    with duckdb.connect() as connection:
        columns = [
            "raw_rows",
            "valid_sale_rows",
            "cancelled_rows",
            "missing_customer_rows",
            "non_positive_quantity_rows",
            "non_positive_price_rows",
            "first_transaction",
            "last_transaction",
            "valid_revenue",
        ]
        row = connection.execute(
            """
            SELECT
                COUNT(*),
                count_if(is_valid_sale),
                count_if(is_cancelled),
                count_if(customer_id IS NULL),
                count_if(quantity <= 0),
                count_if(unit_price <= 0),
                MIN(invoice_timestamp),
                MAX(invoice_timestamp),
                ROUND(SUM(line_revenue) FILTER (WHERE is_valid_sale), 2)
            FROM read_parquet(?)
            """,
            [str(parquet_path)],
        ).fetchone()
    return dict(zip(columns, row))


def build_retail_dataset(
    workbook_path: str | Path,
    output_path: str | Path,
) -> dict[str, Any]:
    """Build a curated Parquet dataset and return its quality report."""
    workbook_path = Path(workbook_path).resolve()
    output_path = Path(output_path).resolve()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Workbook does not exist: {workbook_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    staging_path = output_path.with_suffix(".staging.csv")
    quarantine_path = output_path.with_suffix(".quarantine.csv")

    try:
        source_rows, quarantined_rows = workbook_to_staging_csv(
            workbook_path,
            staging_path,
            quarantine_path,
        )
        _write_curated_parquet(staging_path, output_path)
    finally:
        staging_path.unlink(missing_ok=True)

    if quarantined_rows == 0:
        quarantine_path.unlink(missing_ok=True)

    summary = retail_quality_summary(output_path)
    summary.update(
        {
            "source_rows": source_rows,
            "quarantined_rows": quarantined_rows,
            "source_workbook": workbook_path.name,
            "output_parquet": output_path.name,
        }
    )
    summary_path = output_path.with_suffix(".quality.json")
    summary_path.write_text(
        json.dumps(summary, default=str, indent=2),
        encoding="utf-8",
    )
    return summary


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build the real retail dataset.")
    parser.add_argument("--input", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> None:
    arguments = parse_arguments()
    summary = build_retail_dataset(arguments.input, arguments.output)
    print(json.dumps(summary, default=str, indent=2))


if __name__ == "__main__":
    main()
